from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


TRANSACTION_SHEET_NAME = "交易明细"
LOAN_SUMMARY_SHEET_NAME = "贷款总结"
DEFAULT_FONT_NAME = "微软雅黑"
HEADER_FILL = "000000"
HEADER_FONT_COLOR = "FFFFFF"


def write_workbook(
    transactions: pd.DataFrame,
    summary: pd.DataFrame,
    workbook_file: str | Path,
) -> None:
    """Write the generated transaction and summary dataframes to Excel."""

    workbook_path = Path(workbook_file)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    mode = "a" if workbook_path.exists() else "w"
    writer_kwargs: dict[str, object] = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    try:
        with pd.ExcelWriter(workbook_path, **writer_kwargs) as writer:
            transactions.to_excel(
                writer,
                index=False,
                sheet_name=TRANSACTION_SHEET_NAME,
            )
            summary.to_excel(
                writer,
                index=False,
                sheet_name=LOAN_SUMMARY_SHEET_NAME,
            )
            _format_generated_sheets(
                writer.book,
                [TRANSACTION_SHEET_NAME, LOAN_SUMMARY_SHEET_NAME],
            )
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot update {workbook_path}. Close the workbook and rerun."
        ) from exc


def _format_generated_sheets(workbook, sheet_names: list[str]) -> None:
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        _format_sheet(worksheet)


def _format_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"

    default_font = Font(name=DEFAULT_FONT_NAME)
    header_font = Font(
        name=DEFAULT_FONT_NAME,
        color=HEADER_FONT_COLOR,
        bold=True,
    )
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_FILL,
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = default_font
            cell.alignment = Alignment(vertical="center")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.auto_filter.ref = worksheet.dimensions
